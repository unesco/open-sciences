"""Export API endpoint for downloading search results as XLSX."""

import io
from datetime import datetime

from flask import request, send_file, current_app
from flask.views import MethodView
from invenio_records_resources.services.errors import PermissionDeniedError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportAPIView(MethodView):
    """
    API endpoint to export search results as XLSX.

    URL pattern: /data/export

    Query parameters (same as search page):
        - q (optional): Search query string
        - sort (optional): Sort field
        - p (optional): Page number (default: 1)
        - s (optional): Results per page (default: 10, max: 100)
        - l (optional): Layout (list/grid)
        - All filter parameters from the search page

    Example requests:
        /data/export?q=UNESCO&p=1&s=10
        /data/export?q=climate&resource_type=publication-article
    """

    # Maximum results to export (to prevent server overload)
    MAX_EXPORT_SIZE = 100

    # Column configuration: (header, width, extractor_function)
    COLUMNS = [
        ("Title", 60, lambda r: r.get("metadata", {}).get("title", "")),
        ("Authors", 50, lambda r: _extract_authors(r)),
        ("Affiliations", 60, lambda r: _extract_affiliations(r)),
        ("Resource Type", 20, lambda r: _extract_resource_type(r)),
        ("Publisher", 30, lambda r: r.get("metadata", {}).get("publisher", "")),
        (
            "Publication Date",
            15,
            lambda r: r.get("metadata", {}).get("publication_date", ""),
        ),
        ("Subjects", 50, lambda r: _extract_subjects(r)),
        ("Description", 80, lambda r: r.get("metadata", {}).get("description", "")),
        ("DOI", 30, lambda r: _extract_doi(r)),
        ("Related Identifiers", 50, lambda r: _extract_related_identifiers(r)),
        ("Record URL", 40, lambda r: _extract_record_url(r)),
    ]

    def get(self):
        """Handle GET requests for export."""
        # Get query parameters using InvenioRDM's URL parameter names
        # p = page, s = size (matching search page URL structure)
        query = request.args.get("q", "")
        page = int(request.args.get("p", 1))
        size = min(int(request.args.get("s", 10)), self.MAX_EXPORT_SIZE)
        sort = request.args.get("sort", "newest")

        # Map sort options to service format
        sort_mapping = {
            "newest": "newest",
            "oldest": "oldest",
            "bestmatch": "bestmatch",
            "mostviewed": "mostviewed",
            "mostdownloaded": "mostdownloaded",
            "version": "version",
        }
        sort_option = sort_mapping.get(sort, "newest")

        # Build search parameters
        search_params = {
            "q": query,
            "page": page,
            "size": size,
            "sort": sort_option,
        }

        # Handle facet filters from 'f' parameter
        # Format: f=facet_name:value or f=facet_name:parent+inner:child
        # Convert to InvenioRDM API format: facet_name=value or facet_name=parent::child
        facet_filters = request.args.getlist("f")
        for facet_filter in facet_filters:
            if ":" in facet_filter:
                # Split facet name and value
                facet_name, facet_value = facet_filter.split(":", 1)

                # Handle hierarchical format (e.g., resource_type:publication+inner:publication-article)
                if "+inner:" in facet_value:
                    # Convert parent+inner:child to parent::child format
                    parts = facet_value.split("+inner:")
                    facet_value = "::".join(parts)

                # Add to search params
                # If facet already exists, make it a list
                if facet_name in search_params:
                    existing = search_params[facet_name]
                    if isinstance(existing, list):
                        existing.append(facet_value)
                    else:
                        search_params[facet_name] = [existing, facet_value]
                else:
                    search_params[facet_name] = facet_value

        # Add other filter parameters (excluding known UI parameters)
        excluded_params = {"q", "p", "s", "sort", "l", "f", "page", "size"}
        for key in request.args.keys():
            if key not in excluded_params and key not in search_params:
                values = request.args.getlist(key)
                if values:
                    # If multiple values, pass as list; if single value, pass as string
                    search_params[key] = values if len(values) > 1 else values[0]

        try:
            # Use Flask test client to make internal API request
            # This ensures facet filters are properly applied through the REST API's
            # post_filter mechanism, which the service doesn't support directly
            from urllib.parse import urlencode

            # Build query params for API request
            api_params_list = []
            for key, value in search_params.items():
                if isinstance(value, list):
                    for v in value:
                        api_params_list.append((key, v))
                else:
                    api_params_list.append((key, value))

            # Build URL with properly encoded parameters
            query_string = urlencode(api_params_list)
            api_path = f"/api/records?{query_string}"

            # Make internal request using test client to preserve context
            with current_app.test_client() as client:
                response = client.get(
                    api_path,
                    headers={
                        "Accept": "application/vnd.inveniordm.v1+json",
                        "Host": request.host,
                    },
                )

                if response.status_code != 200:
                    current_app.logger.error(
                        f"Export - API error: {response.status_code} - {response.data}"
                    )
                    return {"error": f"Internal API error: {response.status_code}"}, 500

                results_dict = response.json
                records = results_dict.get("hits", {}).get("hits", [])

            # Generate XLSX
            output = self._generate_xlsx(records, query, page, size)

            # Generate filename
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"unesco_search_export_{date_str}.xlsx"

            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )

        except PermissionDeniedError:
            return {"error": "Permission denied"}, 403
        except Exception as e:
            # Log the error
            current_app.logger.error(f"Export error: {str(e)}", exc_info=True)
            return {"error": str(e)}, 500

    def _generate_xlsx(self, records, query, page, size):
        """Generate XLSX file from records."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="0077D4", end_color="0077D4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (header, width, _) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            for col_idx, (_, _, extractor) in enumerate(self.COLUMNS, start=1):
                try:
                    value = extractor(record)
                except Exception:
                    value = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Add metadata sheet
        meta_ws = wb.create_sheet(title="Export Info")
        meta_data = [
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Search Query", query or "(all records)"),
            ("Page", str(page)),
            ("Results per Page", str(size)),
            ("Total Exported", str(len(records))),
            ("Source", "UNESCO Open Science Portal"),
        ]

        for row_idx, (label, value) in enumerate(meta_data, start=1):
            meta_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            meta_ws.cell(row=row_idx, column=2, value=value)

        meta_ws.column_dimensions["A"].width = 20
        meta_ws.column_dimensions["B"].width = 50

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output


# Helper functions for data extraction


def _extract_authors(record):
    """Extract authors as semicolon-separated string."""
    creators = record.get("metadata", {}).get("creators", [])
    authors = []
    for creator in creators:
        person = creator.get("person_or_org", {})
        name = person.get("name", "")
        if name:
            authors.append(name)
    return "; ".join(authors)


def _extract_affiliations(record):
    """Extract unique affiliations from all creators."""
    creators = record.get("metadata", {}).get("creators", [])
    affiliations = set()
    for creator in creators:
        for aff in creator.get("affiliations", []):
            aff_name = aff.get("name", "") if isinstance(aff, dict) else ""
            if aff_name:
                affiliations.add(aff_name)
    return "; ".join(sorted(affiliations))


def _extract_resource_type(record):
    """Extract resource type title in English."""
    resource_type = record.get("metadata", {}).get("resource_type", {})
    title = resource_type.get("title", {})
    if isinstance(title, dict):
        return title.get("en", title.get("title", ""))
    return str(title) if title else ""


def _extract_subjects(record):
    """Extract subjects as semicolon-separated string."""
    subjects = record.get("metadata", {}).get("subjects", [])
    subject_list = []
    for subj in subjects:
        if isinstance(subj, dict):
            subject_list.append(subj.get("subject", ""))
        elif isinstance(subj, str):
            subject_list.append(subj)
    return "; ".join(filter(None, subject_list))


def _extract_doi(record):
    """Extract DOI from pids or related identifiers."""
    # First try pids
    pids = record.get("pids", {})
    doi_info = pids.get("doi", {})
    if doi_info:
        return doi_info.get("identifier", "")

    # Fallback to related identifiers
    related = record.get("metadata", {}).get("related_identifiers", [])
    for rel in related:
        if rel.get("scheme") == "doi":
            return rel.get("identifier", "")
    return ""


def _extract_related_identifiers(record):
    """Extract related identifiers as formatted string."""
    related = record.get("metadata", {}).get("related_identifiers", [])
    identifiers = []
    for rel in related:
        scheme = rel.get("scheme", "").upper()
        identifier = rel.get("identifier", "")
        relation = rel.get("relation_type", {})
        relation_title = (
            relation.get("title", {}).get("en", "")
            if isinstance(relation, dict)
            else ""
        )
        if scheme and identifier:
            identifiers.append(f"{scheme}: {identifier} ({relation_title})")
    return "; ".join(identifiers)


def _extract_record_url(record):
    """Extract the record URL."""
    record_id = record.get("id", "")
    if record_id:
        # Use relative URL that will work in any environment
        return f"/records/{record_id}"
    return ""
